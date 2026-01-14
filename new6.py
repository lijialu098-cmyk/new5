import streamlit as st
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

# ======================================================
# 主类
# ======================================================
class StreamlitCalculator:
    def __init__(self):
        if "stock_solutions" not in st.session_state:
            self.init_stock_solutions()
        if "molecular_weights" not in st.session_state:
            self.init_mw()
        if "excel_content" not in st.session_state:
            st.session_state.excel_content = None
        if "calculation_success" not in st.session_state:
            st.session_state.calculation_success = False

    # ------------------------
    # 初始化
    # ------------------------
    def init_stock_solutions(self):
        st.session_state.stock_solutions = {
            "Tris": {"concentration": 2.0, "unit": "M", "density": 1.0},
            "NaCl": {"concentration": 5.0, "unit": "M", "density": 1.0},
            "甘油": {"concentration": 100.0, "unit": "%", "density": 1.26},
            "DTT": {"concentration": 1.0, "unit": "M", "density": 1.0},
            "PBS": {"concentration": 10.0, "unit": "X", "density": 1.0},
            "CHAPS": {"concentration": 10.0, "unit": "%", "density": 1.0},
        }

    def init_mw(self):
        st.session_state.molecular_weights = {
            "Tris": 121.14,
            "NaCl": 58.44,
            "甘油": 92.09,
            "DTT": 154.25,
            "CHAPS": 614.88,
        }

    # ------------------------
    # 解析配方 - 修复版本
    # ------------------------
    def parse_formula_string(self, formula_input):
        formula_input = re.sub(r'[，；、]', ',', formula_input)
        
        # 修复正则表达式：正确处理单位
        pattern = r'([\d\.]+)\s*([mMμu]?[Mm]?|[%Xx])\s*([a-zA-Z\u4e00-\u9fa5\-]+)'
        matches = re.findall(pattern, formula_input)
        
        components = {}
        for value, unit, name in matches:
            # 标准化单位
            unit = unit.upper()
            if unit == "MM":
                unit = "mM"
            elif unit == "UM" or unit == "μM":
                unit = "μM"
            
            components[name] = {
                "target_concentration": float(value),
                "target_unit": unit
            }
        
        return components

    # ------------------------
    # 体积解析
    # ------------------------
    def parse_volume(self, text):
        text = text.strip().lower()
        m = re.match(r'([\d\.]+)\s*(l|ml|ul|μl)?', text)
        if not m:
            return None
        v = float(m.group(1))
        unit = m.group(2) or "ml"
        if unit == "l":
            return v * 1000
        if unit in ["ul", "μl"]:
            return v / 1000
        return v

    # ------------------------
    # 计算 - 修复版本
    # ------------------------
    def calculate(self, components, total_ml):
        results = {"components": {}, "total": 0}

        for name, info in components.items():
            target_value = info["target_concentration"]
            target_unit = info["target_unit"]
            
            # 处理库存溶液
            if name in st.session_state.stock_solutions:
                stock = st.session_state.stock_solutions[name]
                stock_unit = stock["unit"]
                
                # 单位转换
                if stock_unit == "M":  # 库存是摩尔浓度
                    if target_unit == "mM":
                        target_value_M = target_value / 1000
                    elif target_unit == "μM":
                        target_value_M = target_value / 1000000
                    elif target_unit == "M":
                        target_value_M = target_value
                    else:
                        st.error(f"不支持的浓度单位: {target_unit}")
                        return None
                    
                    v = (target_value_M * total_ml) / stock["concentration"]
                
                elif stock_unit == "%":  # 库存是百分比
                    if target_unit == "%":
                        v = (target_value * total_ml) / stock["concentration"]
                    else:
                        st.error(f"百分比浓度不匹配: {target_unit}")
                        return None
                
                elif stock_unit == "X":  # 库存是倍数
                    if target_unit == "X":
                        v = (target_value * total_ml) / stock["concentration"]
                    else:
                        st.error(f"倍数不匹配: {target_unit}")
                        return None
                
                results["components"][name] = {
                    "target": f'{target_value} {target_unit}',
                    "volume": v,
                    "mass": v * stock["density"] if "density" in stock else v
                }
                results["total"] += v

            # 处理固体试剂
            elif name in st.session_state.molecular_weights:
                mw = st.session_state.molecular_weights[name]
                
                # 单位转换到摩尔
                if target_unit == "mM":
                    target_mol_per_L = target_value / 1000
                elif target_unit == "μM":
                    target_mol_per_L = target_value / 1000000
                elif target_unit == "M":
                    target_mol_per_L = target_value
                else:
                    st.error(f"不支持的固体浓度单位: {target_unit}")
                    return None
                
                total_L = total_ml / 1000
                mol_needed = target_mol_per_L * total_L
                mass = mol_needed * mw
                
                results["components"][name] = {
                    "target": f'{target_value} {target_unit}',
                    "volume": 0,
                    "mass": mass
                }

        # 计算水的体积
        water = max(0, total_ml - results["total"])
        results["components"]["水"] = {
            "target": "-",
            "volume": water,
            "mass": water
        }
        
        return results

    # ------------------------
    # 写入 Excel - 修复版本（清除多余的水列）
    # ------------------------
    def write_to_excel(self, formula_input, results, total_ml):
        try:
            wb = load_workbook("template.xlsx")
            ws = wb.active

            # 写入基本信息
            ws["C5"] = datetime.now().strftime("%Y-%m-%d")
            ws["C6"] = formula_input
            ws["G6"] = f"{total_ml/1000:.2f} L"

            # 清除之前可能存在的数据（从C列到H列）
            for col in range(3, 9):  # C到H列
                col_letter = get_column_letter(col)
                ws[f"{col_letter}8"] = None  # 清除组分名
                ws[f"{col_letter}11"] = None  # 清除浓度
                ws[f"{col_letter}12"] = None  # 清除质量
                ws[f"{col_letter}13"] = None  # 清除体积

            # 获取所有组分，不包括水
            comps = [(k, v) for k, v in results["components"].items() if k != "水"]
            
            # 写入组分数据
            for i, (name, comp) in enumerate(comps):
                col = get_column_letter(3 + i)  # 从C列开始
                ws[f"{col}8"] = name
                ws[f"{col}11"] = comp["target"]
                if comp["volume"] > 0:
                    ws[f"{col}12"] = round(comp["mass"], 2) if comp["mass"] > 0 else "-"
                    ws[f"{col}13"] = round(comp["volume"], 2)
                else:
                    ws[f"{col}12"] = round(comp["mass"], 4) if comp["mass"] > 0 else "-"
                    ws[f"{col}13"] = "-"

            # 写入水（在最后一个组分之后）
            water_col = get_column_letter(3 + len(comps))
            ws[f"{water_col}8"] = "水"
            ws[f"{water_col}11"] = "-"
            ws[f"{water_col}12"] = round(results["components"]["水"]["mass"], 2)
            ws[f"{water_col}13"] = round(results["components"]["水"]["volume"], 2)

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            st.session_state.excel_content = buffer
            return True

        except Exception as e:
            st.error(f"Excel 生成失败: {e}")
            return False

    # ------------------------
    # UI
    # ------------------------
    def run(self):
        st.title("🧪 试剂配方计算器（模板版）")

        formula = st.text_area(
            "配方输入",
            "20 mM Tris, 150 mM NaCl\n1 mM DTT",
            height=150,
            help="每行或逗号分隔，如: 20 mM Tris, 150 mM NaCl, 1 mM DTT"
        )

        volume = st.text_input("目标体积", "1 L", help="支持 L, mL, μL")

        if st.button("🚀 开始计算", type="primary"):
            # 清空之前的计算结果
            st.session_state.calculation_success = False
            
            total_ml = self.parse_volume(volume)
            if not total_ml:
                st.error("体积格式错误，请使用如: 1 L, 500 mL, 1000 μL")
                return

            comps = self.parse_formula_string(formula)
            
            if not comps:
                st.error("未能解析出有效配方，请检查格式")
                return

            results = self.calculate(comps, total_ml)
            
            if results:
                # 显示计算结果
                st.subheader("📊 计算结果")
                
                # 显示详细结果
                result_df = []
                for name, comp in results["components"].items():
                    if comp["volume"] > 0:
                        result_df.append({
                            "组分": name,
                            "目标浓度": comp["target"],
                            "体积 (mL)": f"{comp['volume']:.2f}",
                            "质量 (g)": f"{comp['mass']:.2f}" if comp["mass"] > 0 else "-"
                        })
                    elif comp["mass"] > 0:
                        result_df.append({
                            "组分": name,
                            "目标浓度": comp["target"],
                            "体积 (mL)": "-",
                            "质量 (g)": f"{comp['mass']:.4f}"
                        })
                    else:
                        result_df.append({
                            "组分": name,
                            "目标浓度": comp["target"],
                            "体积 (mL)": f"{comp['volume']:.2f}",
                            "质量 (g)": f"{comp['mass']:.2f}"
                        })
                
                if result_df:
                    st.dataframe(pd.DataFrame(result_df), use_container_width=True)
                
                # 生成Excel
                ok = self.write_to_excel(formula, results, total_ml)
                if ok:
                    st.session_state.calculation_success = True
                    st.success("✅ 计算完成，可下载 Excel 报告")

        if st.session_state.calculation_success and st.session_state.excel_content:
            st.download_button(
                "📥 下载 Excel 报告",
                st.session_state.excel_content,
                file_name=f"配方计算_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

# ======================================================
# 主入口
# ======================================================
if __name__ == "__main__":
    st.set_page_config(page_title="试剂配方计算器", page_icon="🧪", layout="wide")
    StreamlitCalculator().run()
